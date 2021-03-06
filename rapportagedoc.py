import openpyxl
import os
from datetime import date
from openpyxl.styles import Alignment

class TargetXls:
    #excel van een leerling wordt geopend of aangemaakt
    def __init__(self, klas, naam, nummer, pad):
        self.file = klas + "_" + naam + "_" + str(nummer) + ".xlsx"
        # "D:\\tmp\\voortgang\\leerling\\"
        self.padfile = pad + "\\" + self.file
        self.openofmaakaan()
        self.sheet = self.wb.get_active_sheet()
        sh = self.wb.active
        self.sheet.title = naam
        sh['A1'] = "Voortgangsrapportage excel document van " + naam
        self.wb.save(self.padfile)

    # maak een nieuw bestand of open het bestaande voortgangsbestand
    def openofmaakaan(self):
        try:
            if os.path.isfile(self.padfile):
                self.wb = openpyxl.load_workbook(self.padfile)
            else:
                self.wb = openpyxl.Workbook()
        except IOError as e:
            print("excel could not be created/opened " + e)

    # Maak een nieuwe tab aan, aan het eind en maak het de active worksheet
    def maaktab(self, tabnaam):
        if tabnaam not in self.wb.get_sheet_names():
            self.wb.create_sheet(tabnaam)
        n = len(self.wb.sheetnames)
        self.wb.active = n

    def opendoc(self):
        self.wb = openpyxl.load_workbook(self.padfile)

    def save(self):
        self.wb.save(self.padfile);

    def sluitdoc(self):
        self.wb.close()

    # Paste data from copyRange into template sheet
    def pasteRange(self, startCol, startRow, endCol, endRow, copiedData, sheetName, alignvertical=False):
        if alignvertical:
            al = Alignment(text_rotation=90)
        else:
            al = Alignment(shrinkToFit=True)

        sheetReceiving = self.wb.get_sheet_by_name(sheetName)
        countRow = 0
        for i in range(startRow, endRow + 1, 1):
            countCol = 0
            for j in range(startCol, endCol + 1, 1):
                sheetReceiving.cell(row=i, column=j).value = copiedData[countRow][countCol]
                sheetReceiving.cell(row=i, column=j).alignment = al
                countCol += 1
            countRow += 1

    def zoekvrijerij(self, sheetName):
        week = date.today().isocalendar()[1]
        sheet = self.wb.get_sheet_by_name(sheetName)
        vrijerij = 4
        for i in range(3, 30, 1):
            if i == 3:
                sheet.cell(row=i, column=1).value = "Weken"
                continue
            # De eerste lege rij of als de weekrapportage al geschreven is wordt die rij overschreven
            if (sheet.cell(row=i, column=1).value is None) or (sheet.cell(row=i, column=1).value == "week " + str(week)):
                sheet.cell(row=i, column=1).value = "week " + str(week)
                vrijerij = i
                break
        return vrijerij





