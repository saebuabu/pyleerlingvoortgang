import openpyxl
from tkinter import filedialog
from tkinter import *
import os

class SourceXls:
    def __init__(self):
        # self.file = "IO2A4-voortgang 2017-2018.xlsx"
        vdoc = self.vraagvoortgangsdocument()
        base = os.path.basename(vdoc)
        self.klas = base[0:5]
        # self.pad = "D:\\tmp\\voortgang\\"
        self.wb = openpyxl.load_workbook(vdoc, data_only=True)
        self.sheet = self.wb.get_sheet_by_name(self.klas)
        self.vakken = self.wb.get_sheet_names()
        self.actsheetindex = self.vakken.index(self.klas)
        self.actleerlingindex = -1
        self.leerlingen = []
        self.leerlingDocs = []
        self.namen = []
        self.llcoordinaten = []

    def initsheetindex(self):
        self.actsheetindex = self.vakken.index(self.klas)

    #Leerlingen staan in de tab met de naam van de klas en starten op de 4de rij
    def initleerlingen(self):
        j = 0
        for i in range(1, 50, 1):
            leerling = self.sheet.cell(row=i, column=1).value
            naam = self.sheet.cell(row=i, column=2).value

            if leerling is not None and naam is not None and str(leerling).isdigit():
                # print(str(leerling) + ' ' + str(naam))
                self.leerlingen.append(leerling)
                self.namen.append(naam)
                self.llcoordinaten.append("=" + self.klas + "!"  + self.sheet.cell(row=i, column=2).coordinate)
                j += 1
        #print(self.leerlingen);

    def setnextsheet(self):
        if self.actsheetindex <= len(self.vakken) - 2:
            self.actsheetindex += 1
            self.sheet = self.wb.get_sheet_by_name(self.vakken[self.actsheetindex])
            return True
        else:
            return False

    def getsheetname(self):
        return self.vakken[self.actsheetindex]

    def setnextleerling(self):
        if self.actleerlingindex <= len(self.leerlingen) - 2:
            self.actleerlingindex += 1
            return True
        else:
            return False

    def zoekleerlingrij(self):
        return 4 + self.actleerlingindex

    def kopieerrange(self, startCol, startRow, endCol, endRow):
        sheet = self.sheet
        rangeSelected = []
        # Loops through selected Rows
        for i in range(startRow, endRow + 1, 1):
            # Appends the row to a RowSelected list
            rowSelected = []
            for j in range(startCol, endCol + 1, 1):
                rowSelected.append(sheet.cell(row=i, column=j).value)
            # Adds the RowSelected List and nests inside the rangeSelected
            rangeSelected.append(rowSelected)

        return rangeSelected

    def vraagvoortgangsdocument(self):
        root = Tk()
        root.filename = filedialog.askopenfilename(initialdir="/", title="Selecteer voortgangsdocument")
        return root.filename
